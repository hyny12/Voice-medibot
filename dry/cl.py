



import pandas as pd
from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

def calculate_last_column_average(excel_file):
    # Read Excel file
    df = pd.read_excel(excel_file)
    # Get the last column and calculate its average
    last_column_name = df.columns[-1]
    last_column_average = df[last_column_name].mean()
    return last_column_average

def convert_to_stars(average_rating):
    num_stars = round(average_rating)
    stars = '⭐' * num_stars
    return stars

@app.route('/')
def index():
    return render_template('index2.html')

@app.route('/execute_code', methods=['POST'])
def execute_code():
    import sounddevice as sd
    import speech_recognition as sr
    from textblob import TextBlob
    import tempfile
    import wave

    def speech_to_text():
        recognizer = sr.Recognizer()

        print("Listening...")

        # Record audio using sounddevice
        duration = 3  # seconds
        sample_rate = 44100  # Hz
        audio_data = sd.rec(int(duration * sample_rate), samplerate=sample_rate, channels=1, dtype='int16')
        sd.wait()

        print("Recognizing...")

        # Save audio to a temporary WAV file
        with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
            tmp_file_path = tmp_file.name
            with wave.open(tmp_file_path, 'wb') as wf:
                wf.setnchannels(1)
                wf.setsampwidth(2)
                wf.setframerate(sample_rate)
                wf.writeframes(audio_data.tobytes())

        # Recognize speech using Google Speech Recognition
        try:
            with sr.AudioFile(tmp_file_path) as source:
                audio_data = recognizer.record(source)  # Read the entire audio file
            text = recognizer.recognize_google(audio_data)
            print("Recognized speech:", text)

            # Perform sentiment analysis
            rating, sentiment_score = analyze_sentiment(text)

            print("Rating:", rating)
            print("Sentiment Score:", sentiment_score)

            # Write recognized text and sentiment to a text file
            with open("../recognized_speech_with_sentiment.txt", "w") as file:
                file.write("Recognized speech:\n" + text + "\n\n")
                file.write("Rating: " + str(rating) + "\n")
                file.write("Sentiment Score: " + str(sentiment_score) + "\n")

            return text, rating, sentiment_score

        except sr.UnknownValueError:
            print("Google Speech Recognition could not understand audio")
        except sr.RequestError as e:
            print("Could not request results from Google Speech Recognition service; {0}".format(e))
        finally:
            # Clean up temporary file
            tmp_file.close()
            os.remove(tmp_file_path)  # Remove the temporary WAV file

    def analyze_sentiment(text):
        # Create a TextBlob object
        blob = TextBlob(text)

        # Get the polarity score (-1 to 1)
        sentiment_score = blob.sentiment.polarity

        # Classify sentiment based on keywords
        if 'best' in text.lower():
            sentiment = 5
        elif 'better' in text.lower():
            sentiment = 4
        elif 'good' in text.lower():
            sentiment = 3
        elif 'average' in text.lower():
            sentiment = 3
        elif 'worst' in text.lower():
            sentiment = 1
        elif 'bad' in text.lower():
            sentiment = 1
        else:
            sentiment = 2

        return sentiment, sentiment_score

    def calculate_average(ws):
        # Calculate average rating
        total_ratings = 0
        total_entries = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                if cell.value and str(cell.value).isdigit():
                    total_ratings += cell.value
                    total_entries += 1

        if total_entries > 0:
            average_rating = total_ratings / total_entries
        else:
            average_rating = 0

        return average_rating

    # Paths to the Excel files for each doctor
    doctor1_excel_file_path = "dry/Doc1.xlsx"
    doctor2_excel_file_path = "dry/Doc2.xlsx"
    doctor3_excel_file_path = "dry/Doc3.xlsx"
    doctor4_excel_file_path = "dry/Doc4.xlsx"
    doctor5_excel_file_path = "dry/Doc5.xlsx"
    doctor6_excel_file_path = "dry/Doc6.xlsx"
    doctor7_excel_file_path = "dry/Doc7.xlsx"
    doctor8_excel_file_path = "dry/Doc8.xlsx"
    doctor9_excel_file_path = "dry/Doc9.xlsx"



    # Perform speech recognition
    recognized_text, rating, sentiment_score = speech_to_text()

    if recognized_text:
        print("You said:", recognized_text)

        # Save sentiment analysis to Excel for Doctor 1
        if request.form.get('doctor') == 'Doctor1':
            excel_file_path = doctor1_excel_file_path
        elif request.form.get('doctor') == 'Doctor2':
            excel_file_path = doctor2_excel_file_path
        elif request.form.get('doctor') == 'Doctor3':
            excel_file_path = doctor3_excel_file_path
        elif request.form.get('doctor') == 'Doctor4':
            excel_file_path = doctor4_excel_file_path
        elif request.form.get('doctor') == 'Doctor5':
            excel_file_path = doctor5_excel_file_path
        elif request.form.get('doctor') == 'Doctor6':
            excel_file_path = doctor6_excel_file_path
        elif request.form.get('doctor') == 'Doctor7':
            excel_file_path = doctor7_excel_file_path
        elif request.form.get('doctor') == 'Doctor8':
            excel_file_path = doctor8_excel_file_path
        elif request.form.get('doctor') == 'Doctor9':
            excel_file_path = doctor9_excel_file_path

        try:
            # Load existing workbook or create a new one for the selected doctor
            if os.path.isfile(excel_file_path):
                wb = load_workbook(excel_file_path)
            else:
                wb = Workbook()

            # Get active sheet or create a new one for the selected doctor
            ws = wb.active

            # Check if "Recognized Text" heading already exists for the selected doctor
            if not any(cell.value == "Recognized Text" for row in ws.iter_rows() for cell in row):
                ws.append(["Recognized Text", "Rating", "Sentiment Score"])

            # Append data to Excel file for the selected doctor
            ws.append([recognized_text, rating, sentiment_score])

            # Calculate average rating for the selected doctor
            average_rating = calculate_average(ws)

            # Add or update average rating for the selected doctor
            ws.cell(row=1, column=4, value="Average Rating")
            ws.cell(row=2, column=4, value=average_rating)

            # Save the workbook for the selected doctor
            wb.save(excel_file_path)
            print("Data written to", excel_file_path)

        except FileNotFoundError:
            print("File not found. Creating a new Excel file for the selected doctor...")
            wb = Workbook()
            ws = wb.active
            ws.append(["Recognized Text", "Rating", "Sentiment Score"])
            ws.append([recognized_text, rating, sentiment_score])
            average_rating = calculate_average(ws)
            ws.cell(row=1, column=4, value="Average Rating")
            ws.cell(row=2, column=4, value=average_rating)
            wb.save(excel_file_path)
            print("New Excel file created and data written to", excel_file_path)

    return display_last_column_average()


@app.route('/average_last_column')
def display_last_column_average():
    # Paths to the Excel files for each doctor
    doctor1_excel_file_path = "dry/Doc1.xlsx"
    doctor2_excel_file_path = "dry/Doc2.xlsx"
    doctor3_excel_file_path = "dry/Doc3.xlsx"
    doctor4_excel_file_path = "dry/Doc4.xlsx"
    doctor5_excel_file_path = "dry/Doc5.xlsx"
    doctor6_excel_file_path = "dry/Doc6.xlsx"
    doctor7_excel_file_path = "dry/Doc7.xlsx"
    doctor8_excel_file_path = "dry/Doc8.xlsx"
    doctor9_excel_file_path = "dry/Doc9.xlsx"


    # Calculate average of last column for Doctor 1
    last_column_average_1 = calculate_last_column_average(doctor1_excel_file_path)
    stars_1 = convert_to_stars(last_column_average_1)

    # Calculate average of last column for Doctor 2
    last_column_average_2 = calculate_last_column_average(doctor2_excel_file_path)
    stars_2 = convert_to_stars(last_column_average_2)

    # Calculate average of last column for Doctor 3
    last_column_average_3 = calculate_last_column_average(doctor3_excel_file_path)
    stars_3 = convert_to_stars(last_column_average_3)
    
    # Calculate average of last column for Doctor 4
    last_column_average_4 = calculate_last_column_average(doctor4_excel_file_path)
    stars_4 = convert_to_stars(last_column_average_4)
    
    # Calculate average of last column for Doctor 5
    last_column_average_5 = calculate_last_column_average(doctor5_excel_file_path)
    stars_5 = convert_to_stars(last_column_average_5)
    
    # Calculate average of last column for Doctor 6
    last_column_average_6 = calculate_last_column_average(doctor6_excel_file_path)
    stars_6 = convert_to_stars(last_column_average_6)
    
    # Calculate average of last column for Doctor 7
    last_column_average_7 = calculate_last_column_average(doctor7_excel_file_path)
    stars_7 = convert_to_stars(last_column_average_7)
    
    # Calculate average of last column for Doctor 8
    last_column_average_8 = calculate_last_column_average(doctor8_excel_file_path)
    stars_8 = convert_to_stars(last_column_average_8)
    
    # Calculate average of last column for Doctor 9
    last_column_average_9 = calculate_last_column_average(doctor9_excel_file_path)
    stars_9 = convert_to_stars(last_column_average_9)

    return render_template('index2.html', last_column_average_1=last_column_average_1,display_star_1=stars_1, 
                           last_column_average_2=last_column_average_2, display_star_2=stars_2,
                           last_column_average_3=last_column_average_3, display_star_3=stars_3, 
                           last_column_average_4=last_column_average_4, display_star_4=stars_4, 
                           last_column_average_5=last_column_average_5, display_star_5=stars_5, 
                           last_column_average_6=last_column_average_6, display_star_6=stars_6, 
                           last_column_average_7=last_column_average_7, display_star_7=stars_7, 
                           last_column_average_8=last_column_average_8, display_star_8=stars_8, 
                           last_column_average_9=last_column_average_9, display_star_9=stars_9)


if __name__ == '__main__':
    app.run(debug=True)

