import pandas as pd
from flask import Flask, render_template, request

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/execute_code', methods=['POST'])
def execute_code():
    import sounddevice as sd
    import speech_recognition as sr
    from textblob import TextBlob
    from openpyxl import Workbook, load_workbook
    import tempfile
    import wave
    import os

    def speech_to_text():
        recognizer = sr.Recognizer()

        print("Listening...")

        # Record audio using sounddevice
        duration = 5  # seconds
        sample_rate = 44100  # Hz
        audio_data = sd.rec(int(duration * sample_rate), samplerate=sample_rate, channels=1, dtype='int16')
        sd.wait()

        print("Recognizing...")

        # Save audio to a temporary WAV file
        with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
            tmp_file_path = tmp_file.name
            with wave.open(tmp_file_path, 'wb') as wf:
                wf.setnchannels(1)
                wf.setsampwidth(2)
                wf.setframerate(sample_rate)
                wf.writeframes(audio_data.tobytes())

        # Recognize speech using Google Speech Recognition
        try:
            with sr.AudioFile(tmp_file_path) as source:
                audio_data = recognizer.record(source)  # Read the entire audio file
            text = recognizer.recognize_google(audio_data)
            print("Recognized speech:", text)

            # Perform sentiment analysis
            rating, sentiment_score = analyze_sentiment(text)

            print("Rating:", rating)
            print("Sentiment Score:", sentiment_score)

            # Write recognized text and sentiment to a text file
            with open("../recognized_speech_with_sentiment.txt", "w") as file:
                file.write("Recognized speech:\n" + text + "\n\n")
                file.write("Rating: " + str(rating) + "\n")
                file.write("Sentiment Score: " + str(sentiment_score) + "\n")

            return text, rating, sentiment_score

        except sr.UnknownValueError:
            print("Google Speech Recognition could not understand audio")
        except sr.RequestError as e:
            print("Could not request results from Google Speech Recognition service; {0}".format(e))
        finally:
            # Clean up temporary file
            tmp_file.close()
            os.remove(tmp_file_path)  # Remove the temporary WAV file

    def analyze_sentiment(text):
        # Create a TextBlob object
        blob = TextBlob(text)

        # Get the polarity score (-1 to 1)
        sentiment_score = blob.sentiment.polarity

        # Classify sentiment based on keywords
        if 'best' in text.lower():
            sentiment = 5
        elif 'better' in text.lower():
            sentiment = 4
        elif 'good' in text.lower():
            sentiment = 3
        elif 'average' in text.lower():
            sentiment = 3
        elif 'worst' in text.lower():
            sentiment = 1
        elif 'bad' in text.lower():
            sentiment = 1
        else:
            sentiment = 2

        return sentiment, sentiment_score

    def calculate_average(ws):
        # Calculate average rating
        total_ratings = 0
        total_entries = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                if cell.value and str(cell.value).isdigit():
                    total_ratings += cell.value
                    total_entries += 1

        if total_entries > 0:
            average_rating = total_ratings / total_entries
        else:
            average_rating = 0

        return average_rating

    if __name__ == "__main__":
        # Path to the Excel file
        excel_file_path = "dry/sentiment_analysis.xlsx"
        # Perform speech recognition
        recognized_text, rating, sentiment_score = speech_to_text()

        if recognized_text:
            print("You said:", recognized_text)

            # Save sentiment analysis to Excel
            try:
                # Load existing workbook or create a new one
                if os.path.isfile(excel_file_path):
                    wb = load_workbook(excel_file_path)
                else:
                    wb = Workbook()

                # Get active sheet or create a new one
                ws = wb.active

                # Check if "Recognized Text" heading already exists
                if not any(cell.value == "Recognized Text" for row in ws.iter_rows() for cell in row):
                    ws.append(["Recognized Text", "Rating", "Sentiment Score"])

                # Append data to Excel file
                ws.append([recognized_text, rating, sentiment_score])

                # Calculate average rating
                average_rating = calculate_average(ws)

                # Add or update average rating
                ws.cell(row=1, column=4, value="Average Rating")
                ws.cell(row=2, column=4, value=average_rating)

                # Save the workbook
                wb.save(excel_file_path)
                print("Data written to", excel_file_path)

            except FileNotFoundError:
                print("File not found. Creating a new Excel file...")
                wb = Workbook()
                ws = wb.active
                ws.append(["Recognized Text", "Rating", "Sentiment Score"])
                ws.append([recognized_text, rating, sentiment_score])
                average_rating = calculate_average(ws)
                ws.cell(row=1, column=4, value="Average Rating")
                ws.cell(row=2, column=4, value=average_rating)
                wb.save(excel_file_path)
                print("New Excel file created and data written to", excel_file_path)

    # Place your code here
    # import sounddevice as sd
    # import speech_recognition as sr
    # from textblob import TextBlob
    # from openpyxl import Workbook, load_workbook
    # import tempfile
    # import wave
    # import os
    # Your speech_to_text(), analyze_sentiment(), and calculate_average() functions go here
    
        

    
    return 'Review taken successfully!'

if __name__ == '__main__':
    app.run(debug=True)